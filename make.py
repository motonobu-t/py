from turtle import title
import openpyxl as opxl
import os

from openpyxl import Workbook

class excel:
    #初期化
    def __init__(self, name):
        self.filename = name
        self.isSave = False
        #行を管理
        self.line = 0
        self.row = 1
        self.wb = opxl.Workbook()
        print(self.wb.sheetnames)
        #self.wb.create_sheet(title='Sheet1')
        self.ws = self.wb["Sheet"]

    #対象の文字列を書く
    def write(self, list):

        for fname, count, extr in list:
            #print(fname, count, extr)
            sp = str.split(extr, " : ")
            self.ws.cell(self.row, 1).value = fname
            self.ws.cell(self.row, 2).value = count
            self.ws.cell(self.row, 3).value = sp[0]
            msg : str = sp[1]
            msg = msg.replace('\n', '')
            self.ws.cell(self.row, 4).value = msg
            self.row += 1
        
        self.wb.save(self.filename)
        return

class plantuml:

    component_str =\
        ["CPS",
         "i**UT",
         "i**DT",
         "SPI"]

    msgline = \
        ["->",
        "->>",
        "<-",
        "<<-"]

    idllist = []
    count = 0

    #初期化
    def __init__(self):
        return
        
    def opencheck(self, obj):
        if obj is None:
            return True
        return False
    
    #対象のメッセージ線をチェック
    def linecheck(self, line):
        chk = False
        for l in self.msgline:
            if l in line:
                self.count += 1
                chk = self.check(line)
        return chk
        
    #対象のメッセージなのかチェック
    def check(self, line):
        for compn in self.component_str:
            if compn in line:
                return True
        return False

    #plantumlファイルを読む
    def read(self, obj):
        self.count  = 0
        while True:
            line = obj.readline()
            if self.linecheck(line):
                self.idllist.append([self.fname, self.count, line])
            
            if not line:
                break

    #Plantumlファイルを開く
    def open(self, filename):
        self.fname = filename
        return open(filename, 'r', encoding='utf-8')

    def getlist(self):
        return self.idllist
        
def main():
    filename = 'idllist.xlsx'
    fileexist = os.path.isfile(filename)
    if fileexist == True:
        os.remove(filename)
    
    exc = excel(filename)
    plt = plantuml()
    #mainでplantumlクラスの保存要をチェックして
    #excの保存メソッドをコールする

    for file in os.listdir():
        if '.puml' in file:
            obj = plt.open(file)
            if obj is None:
                continue
            else:
                plt.read(obj)

    list = plt.getlist()
    #listが空かチェック
    if len(list) == 0:
        return
    else:
        #print(list)
        exc.write(list)

    return

if __name__ == '__main__':
    main()
